"""Optional catalog lookup for executing units backed by IT_PEI.xlsx."""

import math
import os
import re
import warnings
from pathlib import Path

# Ocultar la advertencia amarilla de openpyxl
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

try:
    from openpyxl import load_workbook as _load_workbook
except ImportError:  # The server must still start when the optional data file is absent.
    _load_workbook = None

CATALOGO_UE_ENV = 'PEI_IT_PEI_PATH'
# Definimos la ruta de la raíz del proyecto (un nivel arriba de la carpeta backend)
PROJECT_ROOT = Path(__file__).resolve().parent.parent

_ID_UE_PATTERN = re.compile(r'^[0-9]{1,20}$')
_catalogo_cache = None
_catalogo_cache_key = None


class CatalogoUEError(RuntimeError):
    """Raised when the configured Excel catalog cannot be read safely."""


def limpiar_cache_catalogo_ues():
    """Clear the process-local catalog cache, primarily for tests and reloads."""
    global _catalogo_cache, _catalogo_cache_key
    _catalogo_cache = None
    _catalogo_cache_key = None


def resolver_ruta_catalogo(path=None):
    """Resolve only the configured path or the repository's known catalog paths."""
    if path is not None:
        candidate = Path(path).expanduser()
        return candidate if candidate.is_absolute() else PROJECT_ROOT / candidate

    configured = os.environ.get(CATALOGO_UE_ENV)
    if configured:
        candidate = Path(configured).expanduser()
        return candidate if candidate.is_absolute() else PROJECT_ROOT / candidate

    # Apuntamos directamente a la raíz del repositorio donde está el Excel
    candidates = (
        PROJECT_ROOT / 'IT_PEI.xlsx',
    )
    return next((candidate for candidate in candidates if candidate.is_file()), candidates[0])


def normalizar_id_ue(value):
    """Validate the public UE identifier format without interpreting path syntax."""
    text = '' if value is None else str(value)
    if not _ID_UE_PATTERN.fullmatch(text):
        raise ValueError('El ID de UE debe contener entre 1 y 20 digitos.')
    return text


def _id_desde_excel(value):
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None

    text = str(value).strip()
    if re.fullmatch(r'[0-9]+\.0+', text):
        text = text.split('.', 1)[0]
    return text if _ID_UE_PATTERN.fullmatch(text) else None


def _texto(value):
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ''
    return str(value).strip()


def _filas_de_hoja(workbook, sheet_name, period_headers):
    if sheet_name not in workbook.sheetnames:
        raise CatalogoUEError(f'Falta la hoja requerida {sheet_name!r}.')

    rows = workbook[sheet_name].iter_rows(values_only=True)
    try:
        headers = [_texto(value) for value in next(rows)]
    except StopIteration as error:
        raise CatalogoUEError(f'La hoja {sheet_name!r} no tiene encabezados.') from error

    columns = {header: index for index, header in enumerate(headers) if header}
    required = (
        'Id_UE',
        'Nombre_unidad_ejecutora',
        'nombre_provincia',
        'nombre_departamento',
    )
    missing = [header for header in required if header not in columns]
    if missing:
        raise CatalogoUEError(
            f'La hoja {sheet_name!r} no contiene columnas requeridas: {missing}.'
        )

    period_column = next(
        (columns[header] for header in period_headers if header in columns),
        None,
    )
    for row in rows:
        values = list(row)
        id_column = columns['Id_UE']
        id_ue = _id_desde_excel(values[id_column] if id_column < len(values) else None)
        if id_ue is None:
            continue

        def value_for(header):
            index = columns[header]
            return _texto(values[index] if index < len(values) else None)

        yield id_ue, {
            'codigo_ue': id_ue,
            'nombre_municipio': value_for('Nombre_unidad_ejecutora'),
            'nombre_provincia': value_for('nombre_provincia'),
            'nombre_region': value_for('nombre_departamento'),
            'periodo_pei': (
                _texto(values[period_column])
                if period_column is not None and period_column < len(values)
                else ''
            ),
        }


def _leer_catalogo(path):
    if _load_workbook is None:
        raise CatalogoUEError('La dependencia openpyxl no esta disponible.')

    try:
        workbook = _load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            catalogo = {}
            catalogo.update(_filas_de_hoja(workbook, 'IT PEI', ('Periodo PEI',)))
            catalogo.update(_filas_de_hoja(workbook, 'Data_UEs', ('Ult.PEI',)))
            print(f"✅ Catálogo de UEs cargado con éxito: {len(catalogo)} entidades.")
            return catalogo
        finally:
            close = getattr(workbook, 'close', None)
            if close:
                close()
    except CatalogoUEError:
        raise
    except Exception as error:
        raise CatalogoUEError('No se pudo leer el catalogo IT_PEI.xlsx.') from error


def cargar_catalogo_ues(path=None):
    """Load the catalog lazily; an absent file is a valid empty catalog."""
    global _catalogo_cache, _catalogo_cache_key
    resolved_path = resolver_ruta_catalogo(path)
    try:
        stat = resolved_path.stat()
        cache_key = (str(resolved_path), stat.st_mtime_ns, stat.st_size)
    except FileNotFoundError:
        cache_key = (str(resolved_path), None, None)

    if _catalogo_cache_key == cache_key and _catalogo_cache is not None:
        return _catalogo_cache

    if cache_key[1] is None:
        _catalogo_cache = {}
        _catalogo_cache_key = cache_key
        print(f"⚠️ AVISO: No se encontró el catálogo Excel en: {resolved_path}")
        return _catalogo_cache

    _catalogo_cache = _leer_catalogo(resolved_path)
    _catalogo_cache_key = cache_key
    return _catalogo_cache


def obtener_ue(id_ue, catalogo=None):
    """Return only the public fields for one UE, or None when it is absent."""
    normalized_id = normalizar_id_ue(id_ue)
    source = cargar_catalogo_ues() if catalogo is None else catalogo
    return source.get(normalized_id)
